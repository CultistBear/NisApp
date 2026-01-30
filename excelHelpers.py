from openpyxl.styles import Font, Border, Side, Alignment

THICK = Side(style="thick")
THIN = Side(style="thin")

BOLD = Font(bold=True)

LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

NO_BORDER = Border()

def write_cell(ws, row, col, value=None, font=None, align=None, border=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def merge_and_style(ws, r1, c1, r2, c2, value=None, font=None, align=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            write_cell(ws, r, c, font=font, align=align, border=border)

    if value is not None:
        ws.cell(row=r1, column=c1).value = value


def draw_outer_border(ws, start_row, start_col, end_row, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=THICK if c == start_col else cell.border.left,
                right=THICK if c == end_col else cell.border.right,
                top=THICK if r == start_row else cell.border.top,
                bottom=THICK if r == end_row else cell.border.bottom,
            )
