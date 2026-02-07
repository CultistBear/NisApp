from openpyxl import Workbook
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter
from datetime import date

from excelHelpers import (
    write_cell,
    merge_and_style,
    draw_outer_border,
    BOLD,
    LEFT,
    RIGHT,
    LEFT_WRAP,
    THICK,
    CENTER
)
from constants import INPUT_TYPE_SUBTYPES


def render_subtype_table(ws, start_row, start_col, type_name, subtype_name, columns, rows, fixed_height):
    current_row = start_row
    width = len(columns)

    merge_and_style(
        ws,
        current_row,
        start_col,
        current_row,
        start_col + width - 1,
        f"{type_name} - {subtype_name}",
        font=BOLD,
        align=LEFT,
        border=Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
    )
    current_row += 1

    for i, col in enumerate(columns):
        write_cell(
            ws,
            current_row,
            start_col + i,
            col["name"],
            font=BOLD,
            align=LEFT,
            border=Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
        )
    current_row += 1

    data_start_row = current_row
    for row in rows:
        for i, value in enumerate(row):
            if i > 0:
                write_cell(ws, current_row, start_col + i, value, align=LEFT_WRAP)
            else:
                write_cell(ws, current_row, start_col + i, value, align=CENTER)
        current_row += 1
    
    rows_written = len(rows)
    for _ in range(fixed_height - rows_written):
        current_row += 1
    
    data_end_row = data_start_row + len(rows) - 1

    amount_col = start_col
    amount_letter = get_column_letter(amount_col)

    if len(rows) > 0:
        formula = f"=SUM({amount_letter}{data_start_row}:{amount_letter}{data_end_row})"
    else:
        formula = 0

    write_cell(
        ws,
        current_row,
        amount_col,
        formula,
        font=BOLD,
        align=RIGHT,
        border=Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
    )

    write_cell(
        ws,
        current_row,
        amount_col + 1,
        "SUBTOTAL",
        font=BOLD,
        align=LEFT
    )

    subtotal_cell_ref = f"{amount_letter}{current_row}"

    draw_outer_border(ws, start_row, start_col, current_row, start_col + width - 1)

    return subtotal_cell_ref


def render_empty_subtype_slot(ws, start_row, start_col, width, fixed_height):
    return start_row + 1 + 1 + fixed_height


def generate_excel(column_map, db_data, report_date=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Register"

    INPUT_TYPES = list(INPUT_TYPE_SUBTYPES.keys())

    for col in range(1, 50):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.column_dimensions['A'].width = 3

    if report_date is None:
        report_date = date.today().strftime('%d-%m-%Y')

    merge_and_style(
        ws,
        1,
        2,
        1,
        15,
        f"Expense Register — {report_date}",
        font=BOLD,
        align=LEFT
    )

    START_ROW = 3
    START_COL = 2
    COL_GAP = 1
    GAP_WIDTH = 3

    types_list = list(column_map.keys())
    
    all_subtypes = set()
    for subtypes in INPUT_TYPE_SUBTYPES.values():
        all_subtypes.update(subtypes)
    
    subtype_max_rows = {}
    for subtype_name in all_subtypes:
        max_rows = 1
        for type_name in types_list:
            for input_type in INPUT_TYPES:
                rows = db_data.get(type_name, {}).get(input_type, {}).get(subtype_name, [])
                max_rows = max(max_rows, len(rows))
        subtype_max_rows[subtype_name] = max_rows

    input_type_subtype_start_rows = {}
    for input_type in INPUT_TYPES:
        input_type_subtype_start_rows[input_type] = {}
        current_row = START_ROW
        for subtype_name in INPUT_TYPE_SUBTYPES[input_type]:
            input_type_subtype_start_rows[input_type][subtype_name] = current_row
            table_height = 1 + 1 + subtype_max_rows[subtype_name] + 1 
            current_row += table_height + 1  

    input_type_heights = {}
    for input_type in INPUT_TYPES:
        total_height = 0
        for subtype_name in INPUT_TYPE_SUBTYPES[input_type]:
            table_height = 1 + 1 + subtype_max_rows[subtype_name] + 1 + 1
            total_height += table_height
        input_type_heights[input_type] = total_height

    max_column_height = max(input_type_heights.values())
    earnings_totals_row = START_ROW + max_column_height + 1
    payments_totals_row = earnings_totals_row + 2  

    type_columns = {}  
    type_widths = {}
    gap_columns = []
    current_col = START_COL

    for idx, type_name in enumerate(types_list):
        type_columns[type_name] = {}
        
        max_width = 2
        if type_name in column_map:
            for subtype_name, columns in column_map[type_name].items():
                max_width = max(max_width, len(columns))
        type_widths[type_name] = max_width

        for input_type in INPUT_TYPES:
            type_columns[type_name][input_type] = current_col
            current_col += max_width
        
        if idx < len(types_list) - 1:
            gap_columns.append(current_col)
            current_col += COL_GAP

    for gap_col in gap_columns:
        ws.column_dimensions[get_column_letter(gap_col)].width = GAP_WIDTH

    type_subtotal_cells = {type_name: {it: [] for it in INPUT_TYPES} for type_name in types_list}
    subtype_subtotal_cells = {subtype_name: [] for subtype_name in all_subtypes}
    input_type_subtotal_cells = {it: [] for it in INPUT_TYPES}

    for type_name in types_list:
        width = type_widths[type_name]
        
        for input_type in INPUT_TYPES:
            col = type_columns[type_name][input_type]
            
            for subtype_name in INPUT_TYPE_SUBTYPES[input_type]:
                row = input_type_subtype_start_rows[input_type][subtype_name]
                fixed_height = subtype_max_rows[subtype_name]
                
                if type_name in column_map and subtype_name in column_map[type_name]:
                    columns = column_map[type_name][subtype_name]
                    columns = sorted(columns, key=lambda c: c["name"] != "Amount(₹)")
                    rows = db_data.get(type_name, {}).get(input_type, {}).get(subtype_name, [])
                    
                    subtotal_ref = render_subtype_table(
                        ws,
                        row,
                        col,
                        type_name,
                        subtype_name,
                        columns,
                        rows,
                        fixed_height
                    )
                    type_subtotal_cells[type_name][input_type].append(subtotal_ref)
                    subtype_subtotal_cells[subtype_name].append(subtotal_ref)
                    input_type_subtotal_cells[input_type].append(subtotal_ref)

    earnings_total_refs = []
    payments_total_refs = []
    
    for type_name in types_list:
        for input_type in INPUT_TYPES:
            col = type_columns[type_name][input_type]
            amount_letter = get_column_letter(col)
            
            subtotal_cells = type_subtotal_cells[type_name][input_type]
            if subtotal_cells:
                type_formula = f"=SUM({','.join(subtotal_cells)})"
            else:
                type_formula = 0

            if input_type == "EARNINGS":
                target_row = earnings_totals_row
                earnings_total_refs.append(f"{amount_letter}{target_row}")
            else:
                target_row = payments_totals_row
                payments_total_refs.append(f"{amount_letter}{target_row}")

            write_cell(
                ws,
                target_row,
                col,
                type_formula,
                font=BOLD,
                align=RIGHT,
                border=Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
            )

            write_cell(
                ws,
                target_row,
                col + 1,
                f"{type_name} {input_type}",
                font=BOLD,
                align=LEFT
            )

    subtype_totals_col = current_col + 1
    payments_subtype_totals_col = subtype_totals_col + 3
    
    for input_type in INPUT_TYPES:
        for subtype_name in INPUT_TYPE_SUBTYPES[input_type]:
            subtype_row = input_type_subtype_start_rows[input_type][subtype_name]
            subtotal_row = subtype_row + 1 + 1 + subtype_max_rows[subtype_name]
            
            if input_type == "EARNINGS":
                col_for_total = subtype_totals_col
            else:
                col_for_total = payments_subtype_totals_col
            
            amount_letter = get_column_letter(col_for_total)
            
            subtotal_cells = subtype_subtotal_cells[subtype_name]
            if subtotal_cells:
                subtype_formula = f"=SUM({','.join(subtotal_cells)})"
            else:
                subtype_formula = 0

            write_cell(
                ws,
                subtotal_row,
                col_for_total,
                subtype_formula,
                font=BOLD,
                align=RIGHT,
                border=Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
            )

            write_cell(
                ws,
                subtotal_row,
                col_for_total + 1,
                f"{subtype_name} TOTAL",
                font=BOLD,
                align=LEFT
            )

    total_col = subtype_totals_col
    total_letter = get_column_letter(total_col)

    if earnings_total_refs:
        earnings_formula = f"=SUM({','.join(earnings_total_refs)})"
    else:
        earnings_formula = 0

    write_cell(
        ws,
        earnings_totals_row,
        total_col,
        earnings_formula,
        font=BOLD,
        align=RIGHT,
        border=Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
    )

    write_cell(
        ws,
        earnings_totals_row,
        total_col + 1,
        "EARNINGS TOTAL",
        font=BOLD,
        align=LEFT
    )

    payments_total_col = payments_subtype_totals_col
    if payments_total_refs:
        payments_formula = f"=SUM({','.join(payments_total_refs)})"
    else:
        payments_formula = 0

    write_cell(
        ws,
        payments_totals_row,
        payments_total_col,
        payments_formula,
        font=BOLD,
        align=RIGHT,
        border=Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
    )

    write_cell(
        ws,
        payments_totals_row,
        payments_total_col + 1,
        "PAYMENTS TOTAL",
        font=BOLD,
        align=LEFT
    )

    max_data_col = payments_subtype_totals_col + 2
    used_cols = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_data_col):
        for cell in row:
            if cell.value is not None and cell.value != "":
                used_cols.add(cell.column)
    
    for col in range(1, max_data_col + 1):
        if col not in used_cols:
            ws.column_dimensions[get_column_letter(col)].width = 3

    return wb
